from __future__ import annotations
import io, re
from pathlib import Path
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.cell import get_column_letter

PLACEHOLDER = re.compile(r"\{\{\s*([^\}]+?)\s*\}\}")

# ---------- 基本ユーティリティ ----------
def _pick(d: Dict[str, Any], dotted: str, default: Any = "") -> Any:
    cur: Any = d
    for k in dotted.split("."):
        if isinstance(cur, dict) and k in cur:
            cur = cur[k]
        else:
            return default
    return cur

def _format_value(key: str, val: Any) -> Any:
    if val is None:
        return ""
    if isinstance(val, bool):
        return "有" if val else "無"
    if key.endswith("_date") and isinstance(val, str):
        try:
            return datetime.fromisoformat(val).date().strftime("%Y-%m-%d")
        except ValueError:
            return val
    return val

def _replace_placeholders(ws, ctx: Dict[str, Any]) -> None:
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = cell.value
            if not isinstance(v, str) or "{{" not in v:
                continue

            def repl(m: re.Match) -> str:
                key = m.group(1).strip()
                raw = _pick(ctx, key, None)
                return str(_format_value(key, raw)) if raw is not None else ""

            new_v = PLACEHOLDER.sub(repl, v)
            if new_v != v:
                cell.value = new_v

# ---------- items 正規化 ----------
def _to_number(x: Any) -> float | int:
    if x is None:
        return 0
    if isinstance(x, (int, float)):
        return x
    if isinstance(x, str):
        s = x.strip().replace(",", "")
        try:
            if "." in s:
                return float(s)
            return int(s)
        except ValueError:
            return 0
    return 0

def _normalize_item(it: Any) -> Optional[Dict[str, Any]]:
    """
    items[*] が dict 以外（str など）でも落ちないように正規化。
    既存スキーマの差異もここで吸収する。
    """
    if isinstance(it, str):
        return {
            "description": it,
            "qty": 0,
            "unit_price": 0,
            "notes": "",
        }
    if not isinstance(it, dict):
        return None

    # description
    desc = it.get("description")
    if not desc:
        dev = it.get("device_name") or ""
        lay = it.get("layer_name") or ""
        desc = f"{dev} / {lay}".strip(" /")

    # qty
    qty = it.get("qty")
    if qty is None:
        qty = it.get("quantity")

    # unit price
    unit = it.get("unit_price")
    if unit is None:
        unit = it.get("unit_price_jpy")

    # notes（備考）
    notes = it.get("notes")
    if not notes:
        rids = it.get("reticle_ids")
        if isinstance(rids, list):
            notes = ", ".join(map(str, rids))
    if not notes:
        notes = it.get("sec_invoice_no")

    return {
        "description": desc or "",
        "qty": _to_number(qty),
        "unit_price": _to_number(unit),
        "notes": "" if notes is None else str(notes),
    }

# ---------- items シート出力 ----------
def _insert_items_sheet(wb, items: List[Any], sheet_name: str = "別紙") -> None:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    headers = ["品名", "数量", "単価", "金額", "備考"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=1, column=j, value=h).alignment = Alignment(horizontal="center")

    row = 2
    for raw in items or []:
        it = _normalize_item(raw)
        if not it:
            continue

        desc = it["description"]
        qty = _to_number(it.get("qty"))
        unit = _to_number(it.get("unit_price"))
        amount = qty * unit
        notes = it.get("notes") or ""

        ws.cell(row=row, column=1, value=desc)
        c = ws.cell(row=row, column=2, value=qty);  c.number_format = "#,##0"
        c = ws.cell(row=row, column=3, value=unit); c.number_format = "#,##0"
        c = ws.cell(row=row, column=4, value=amount); c.number_format = "#,##0"
        ws.cell(row=row, column=5, value=notes)
        row += 1

    n = (row - 2)
    if n > 0:
        ws.cell(row=n + 2, column=3, value="小計")
        c = ws.cell(row=n + 2, column=4, value=f"=SUM(D2:D{n+1})"); c.number_format = "#,##0"

    # 幅調整
    for col in range(1, 6):
        max_len = 0
        for r in range(1, n + 3):
            val = ws.cell(row=r, column=col).value
            max_len = max(max_len, len(str(val or "")))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 24)

# ---------- コンテキスト作成とレンダ ----------
def _shallow_merge(dst: Dict[str, Any], src: Dict[str, Any]) -> None:
    for k, v in src.items():
        if k not in dst:
            dst[k] = v

def _detect_items(data: Dict[str, Any]) -> List[Any]:
    for path in [
        ("items",),
        ("purchase_order", "items"),
        ("non_exemption_certificate", "items"),
    ]:
        cur: Any = data
        ok = True
        for k in path:
            if isinstance(cur, dict) and k in cur:
                cur = cur[k]
            else:
                ok = False
                break
        if ok and isinstance(cur, list):
            return cur
    return []

def render_excel_from_json(data: Dict[str, Any], template_path: Path, detail_sheet_name: str = "別紙") -> bytes:
    # ネスト保持＋平坦化
    ctx: Dict[str, Any] = dict(data)

    po = data.get("purchase_order") or {}
    if isinstance(po, dict):
        _shallow_merge(ctx, po)
        for subkey in ("order_details", "buyer", "seller"):
            sub = po.get(subkey)
            if isinstance(sub, dict):
                _shallow_merge(ctx, sub)

    nec = data.get("non_exemption_certificate")
    if isinstance(nec, dict):
        _shallow_merge(ctx, nec)

    wb = load_workbook(template_path)
    for ws in wb.worksheets:
        _replace_placeholders(ws, ctx)

    items = _detect_items(data)
    _insert_items_sheet(wb, items, sheet_name=detail_sheet_name)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
